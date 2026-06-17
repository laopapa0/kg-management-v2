import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

xlsx_path = r"C:\Users\pan47\Desktop\Demo v2\功能清单_React版_Vue重构基准.xlsx"

# ── 估算数据 ──
estimate_data = [
    # 业务场景, 预计AI对话轮数, 复杂度, 风险说明

    # ── 基础设施（新增sheet） ──
    ["Vue 项目脚手架 + 路由 + 菜单 + 布局", "2-3", "🟡 中", "Vite + Vue Router + Layout 框架搭建"],
    ["深色主题系统 + Pinia 状态管理", "3-4", "🟡 中", "CSS 变量体系迁移，Zustand → Pinia 重写"],
    ["UI 组件库集成 + 通用组件移植", "2-3", "🟢 低", "Element Plus / Naive UI 替换 shadcn/ui"],

    # ── 首页/工作台 ──
    ["概览统计卡片", "1-2", "🟢 低", "纯展示组件，无复杂交互"],
    ["快捷入口导航", "1", "🟢 低", "简单路由跳转卡片"],
    ["近期动态时间线", "1-2", "🟢 低", "列表渲染 + 图标映射"],

    # ── 指标管理 ──
    ["指标挂靠", "10-14", "🔴 高", "最复杂场景：SVG 实时连线跟随鼠标、hover 检测、动画反馈、持久连线层、脑图投递区交互、多目标类型识别"],
    ["指标树高级交互", "6-8", "🟡 中", "树组件重构、inline 编辑、拖拽排序、搜索定位，可用 Element Plus Tree 降本"],
    ["脑图可视化+交互", "8-12", "🔴 高", "Mind Elixir 5 无 Vue3 官方封装，需手写 wrapper；主题映射、事件同步、连线投递联动"],
    ["标签树高级交互", "3-4", "🟢 低", "树 + Space 切换选中，较简单"],
    ["规则树高级交互", "3-4", "🟢 低", "树 + 参数侧边栏联动"],
    ["指标配置", "4-6", "🟡 中", "命令面板 + 侧边栏联动 + 批量关联"],

    # ── 血缘画布 ──
    ["血缘关系可视化+实例管理", "8-12", "🔴 高", "SVG 画布自绘：节点布局/连线/拖拽/缩放/高亮路径，无现成组件可复用"],

    # ── 报告管理 ──
    ["报告计划管理", "3-4", "🟢 低", "CRUD 表单 + 列表 + 开关"],
    ["报告生成与历史", "4-6", "🟡 中", "列表/筛选/分页/版本回溯、新窗口打开报告"],
    ["报告模板管理", "4-6", "🟡 中", "CRUD + 章节拖拽排序 + AI 生成交互"],

    # ── 规则管理 ──
    ["规则树高级交互", "3-4", "🟢 低", "树展开折叠 + 拖拽排序"],
    ["规则实例增删改查", "4-6", "🟡 中", "列表 + JSON 编辑器 + 父规则继承"],

    # ── 关联关系管理 ──
    ["关联关系实例管理", "4-6", "🟡 中", "表格 CRUD + 展开详情 + 变更时间线"],
    ["AI 推荐关系管理", "3-5", "🟢 低", "列表筛选 + 勾选应用"],

    # ── 知识库管理 ──
    ["知识库管理", "1-2", "🟢 低", "纯后端 API 对接，前端仅文档列表/上传"],
]

# ── 分类 ──
infra = estimate_data[:3]       # 基础设施
business = estimate_data[3:]     # 业务场景

# ── 读已有 xlsx ──
wb = load_workbook(xlsx_path)

# ── 样式 ──
header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

body_font = Font(name="Microsoft YaHei", size=10)
body_align = Alignment(vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

red_fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
yellow_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
green_fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
infra_fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")

# ── 新建 Sheet2: 工作量估算 ──
ws = wb.create_sheet("AI重构工作量估算")

headers = ["分类", "业务场景 / 基础设施", "预计 AI 对话轮数", "复杂度", "风险说明"]
for ci, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = header_align
    c.border = thin_border

# 写基础设施行
for ri, row in enumerate(infra, 2):
    category = "基础设施"
    vals = [category] + list(row)
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=ri, column=ci, value=v)
        c.font = body_font
        c.alignment = body_align
        c.border = thin_border
        if ci == 1:
            c.fill = infra_fill

# 分隔行
sep_row = len(infra) + 2
ws.merge_cells(start_row=sep_row, start_column=1, end_row=sep_row, end_column=5)
sc = ws.cell(row=sep_row, column=1, value="── 业务功能 ──")
sc.font = Font(name="Microsoft YaHei", size=10, bold=True, color="6B7280")
sc.alignment = Alignment(horizontal="center", vertical="center")

# 写业务行
for ri, row in enumerate(business, sep_row + 1):
    category = "业务功能"
    vals = [category] + list(row)
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=ri, column=ci, value=v)
        c.font = body_font
        c.alignment = body_align
        c.border = thin_border
        if ci == 1:
            c.fill = infra_fill
        # 复杂度列着色
        if ci == 4:
            if "🔴" in str(v):
                c.fill = red_fill
            elif "🟡" in str(v):
                c.fill = yellow_fill
            elif "🟢" in str(v):
                c.fill = green_fill

# ── 汇总行 ──
total_row = sep_row + 1 + len(business) + 2
ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
tc = ws.cell(row=total_row, column=1, value="合计")
tc.font = Font(name="Microsoft YaHei", size=11, bold=True, color="1F2937")
tc.alignment = Alignment(horizontal="right", vertical="center")

tc2 = ws.cell(row=total_row, column=2, value="基础设施 7-10 轮 + 业务场景 69-100 轮")
tc2.font = Font(name="Microsoft YaHei", size=11, bold=True, color="1F2937")
tc2.alignment = body_align

total_range = f"{76}-{110}"
tc3 = ws.cell(row=total_row, column=3, value=total_range)
tc3.font = Font(name="Microsoft YaHei", size=11, bold=True, color="DC2626")
tc3.alignment = body_align

tc4 = ws.cell(row=total_row, column=4, value="约 4-6 周（1 人全时）")
tc4.font = Font(name="Microsoft YaHei", size=11, bold=True, color="1F2937")
tc4.alignment = body_align

# ── 列宽 ──
ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 26
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 10
ws.column_dimensions["E"].width = 52

# 行高
for r in range(1, total_row + 1):
    ws.row_dimensions[r].height = 22

ws.row_dimensions[1].height = 28
ws.row_dimensions[sep_row].height = 24
ws.row_dimensions[total_row].height = 28

# 冻结
ws.freeze_panes = "A2"

wb.save(xlsx_path)
print(f"已更新: {xlsx_path}")
print(f"预估 AI 对话轮数: {total_range} 轮")
