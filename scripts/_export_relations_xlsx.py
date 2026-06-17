import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def extract_field(obj, field):
    m = re.search(rf'{field}:\s*"([^"]*)"', obj)
    return m.group(1) if m else ''

# 读取 indicatorDefinitions
with open('src/data/indicatorDefinitions.ts', 'r', encoding='utf-8') as f:
    ind = f.read()

blocks = re.findall(r'\{[^}]+\}', ind)
info_map = {}
for b in blocks:
    code = extract_field(b, 'code')
    if not code:
        continue
    info_map[code] = {
        'name': extract_field(b, 'name'),
        'level1': extract_field(b, 'level1'),
        'level2': extract_field(b, 'level2'),
        'department': extract_field(b, 'department'),
        'frequency': extract_field(b, 'frequency'),
        'granularity': extract_field(b, 'granularity'),
    }

# 读取 aiRecommendations
with open('src/data/aiRecommendations.ts', 'r', encoding='utf-8') as f:
    ai = f.read()

def parse_connections(arr_text):
    result = []
    blocks = re.findall(r'\{[^}]+\}', arr_text)
    for b in blocks:
        src = extract_field(b, 'sourceId')
        tgt = extract_field(b, 'targetId')
        typ = extract_field(b, 'relationTypeId')
        conf = extract_field(b, 'confidence')
        reason = extract_field(b, 'reason')
        if not src or not tgt:
            continue
        result.append({
            'sourceId': src,
            'targetId': tgt,
            'relationTypeId': typ,
            'confidence': conf or '',
            'reason': reason or '',
        })
    return result

# 提取两个数组
start1 = ai.find('export const mockAppliedConnections')
end1 = ai.find('];', start1)
applied = parse_connections(ai[start1:end1 + 2])

start2 = ai.find('export const mockAiRecommendations')
end2 = ai.find('];', start2)
unapplied = parse_connections(ai[start2:end2 + 2])

print(f'已应用: {len(applied)}, 未应用: {len(unapplied)}')

# 创建 Excel
wb = openpyxl.Workbook()

header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2B579A', end_color='2B579A', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_align = Alignment(vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

# Sheet 1: 已应用
ws1 = wb.active
ws1.title = '已应用-血缘关系'

headers = ['#', '状态', 'sourceId', '源指标名称', '源level1', '源level2', '源部门',
           'targetId', '目标指标名称', '目标level1', '目标level2', '目标部门',
           'relationTypeId', 'frequency', 'granularity']
for c, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

for i, r in enumerate(applied):
    si = info_map.get(r['sourceId'], {})
    ti = info_map.get(r['targetId'], {})
    row = [
        i + 1, '已应用',
        r['sourceId'], si.get('name', ''), si.get('level1', ''), si.get('level2', ''),
        si.get('department', ''),
        r['targetId'], ti.get('name', ''), ti.get('level1', ''), ti.get('level2', ''),
        ti.get('department', ''),
        r['relationTypeId'], si.get('frequency', ''), si.get('granularity', ''),
    ]
    for c, v in enumerate(row, 1):
        cell = ws1.cell(row=i + 2, column=c, value=v)
        cell.alignment = cell_align
        cell.border = thin_border

# Sheet 2: 未应用
ws2 = wb.create_sheet('未应用-AI推荐')

headers2 = ['#', '状态', 'sourceId', '源指标名称', '源level1', '源level2', '源部门',
            'targetId', '目标指标名称', '目标level1', '目标level2', '目标部门',
            'relationTypeId', 'confidence', 'reason']
for c, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color='D4A017', end_color='D4A017', fill_type='solid')
    cell.alignment = header_align
    cell.border = thin_border

for i, r in enumerate(unapplied):
    si = info_map.get(r['sourceId'], {})
    ti = info_map.get(r['targetId'], {})
    row = [
        i + 1, '未应用',
        r['sourceId'], si.get('name', ''), si.get('level1', ''), si.get('level2', ''),
        si.get('department', ''),
        r['targetId'], ti.get('name', ''), ti.get('level1', ''), ti.get('level2', ''),
        ti.get('department', ''),
        r['relationTypeId'], r['confidence'], r['reason'],
    ]
    for c, v in enumerate(row, 1):
        cell = ws2.cell(row=i + 2, column=c, value=v)
        cell.alignment = cell_align
        cell.border = thin_border

# Sheet 3: 财务部（已应用 + 未应用）
ws3 = wb.create_sheet('财务部-全部')

headers3 = ['#', '状态', 'sourceId', '源指标名称', '源level1', '源level2',
            'targetId', '目标指标名称', '目标level1', '目标level2',
            'relationTypeId', 'confidence', 'reason']
for c, h in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=c, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    cell.alignment = header_align
    cell.border = thin_border

fin_lines = []
all_records = [('已应用', r) for r in applied] + [('未应用', r) for r in unapplied]
for status, r in all_records:
    si = info_map.get(r['sourceId'], {})
    ti = info_map.get(r['targetId'], {})
    if si.get('department') == '财务部' and ti.get('department') == '财务部':
        fin_lines.append({
            'status': status,
            'sourceId': r['sourceId'],
            'sourceName': si.get('name', ''),
            'sourceLevel1': si.get('level1', ''),
            'sourceLevel2': si.get('level2', ''),
            'targetId': r['targetId'],
            'targetName': ti.get('name', ''),
            'targetLevel1': ti.get('level1', ''),
            'targetLevel2': ti.get('level2', ''),
            'relationTypeId': r['relationTypeId'],
            'confidence': r['confidence'],
            'reason': r['reason'],
        })

for i, r in enumerate(fin_lines):
    row = [
        i + 1, r['status'],
        r['sourceId'], r['sourceName'], r['sourceLevel1'], r['sourceLevel2'],
        r['targetId'], r['targetName'], r['targetLevel1'], r['targetLevel2'],
        r['relationTypeId'], r['confidence'], r['reason'],
    ]
    for c, v in enumerate(row, 1):
        cell = ws3.cell(row=i + 2, column=c, value=v)
        cell.alignment = cell_align
        cell.border = thin_border

# 调整列宽
for ws in [ws1, ws2, ws3]:
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 36
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 22
    ws.column_dimensions['I'].width = 36
    ws.column_dimensions['J'].width = 12

out_path = '.handoff/关联关系实例全量导出.xlsx'
wb.save(out_path)
print(f'导出完成: {out_path}')
print(f'  已应用: {len(applied)} 条')
print(f'  未应用: {len(unapplied)} 条')
print(f'  财务部: {len(fin_lines)} 条')
