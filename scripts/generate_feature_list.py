import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

out_dir = os.path.join(os.path.dirname(__file__), "..")
out_path = os.path.join(out_dir, "feature_list.xlsx")

# ========== Sheet 1: 功能清单 ==========

data = [
    # 模块, 业务场景, 描述, 子功能拆解(前端/后端+完成状态), 核心files

    # ── 首页 ──
    ["首页/工作台", "概览统计卡片",
     "展示项目数据概要（已接入指标、血缘关系数、活跃规则数等）",
     "【前端】 1. 统计卡片展示 ✅\n【前端】 2. 趋势变化率显示 ✅",
     "src/pages/dashboard/DashboardPage.tsx"],
    ["首页/工作台", "快捷入口导航",
     "快速跳转四大模块（指标管理、血缘画布、报告管理、知识库管理）",
     "【前端】 1. 快捷卡片渲染 ✅\n【前端】 2. 路由跳转 ✅",
     "src/pages/dashboard/DashboardPage.tsx"],
    ["首页/工作台", "近期动态时间线",
     "展示操作记录时间线（创建/变更/告警等动态）",
     "【前端】 1. 时间线列表 ✅\n【前端】 2. 操作类型图标标识 ✅",
     "src/pages/dashboard/DashboardPage.tsx"],

    # ── 指标管理 ──
    ["指标管理", "指标挂靠",
     "待选指标通过连线方式挂靠到指标树/标签集/规则节点",
     "【前端】 1. 连线模式（Space 进入\u2192SVG 虚线跟随鼠标\u2192ESC/右键取消）✅\n"
     "【前端】 2. 目标区域 hover 检测与高亮 ✅\n"
     "【前端】 3. 点击目标完成挂靠（树节点/脑图默认区）✅\n"
     "【前端】 4. 脉冲+反馈动画 ✅\n"
     "【前端】 5. 挂靠关系持久连线展示 ✅\n"
     "【前端】 6. 删除连线 + 撤销 ✅\n"
     "【前端】 7. 误操作抖动提示 ✅\n"
     "【后端】 1. 挂靠关系持久化（当前 localStorage）✅",
     "IndicatorAttachmentPage.tsx\nConnectionLayer.tsx\nPersistentConnectionLayer.tsx\nattachmentStore.ts"],
    ["指标管理", "指标树高级交互",
     "指标树节点层级操作与可视化交互",
     "【前端】 1. 节点展开/折叠 ✅\n"
     "【前端】 2. 节点 inline 重命名（双击）✅\n"
     "【前端】 3. 节点新增与删除 ✅\n"
     "【前端】 4. 拖拽调整层级归属 ✅\n"
     "【后端】 1. 指标树节点 CRUD API \U0001F534\n"
     "【后端】 2. 指标树层级关系更新 API \U0001F534",
     "IndicatorTreePanel.tsx\nTreeView.tsx\nTreeNodeInlineEdit.tsx\nattachmentTree.ts"],
    ["指标管理", "脑图可视化+交互",
     "指标树思维导图可视化（Mind Elixir 5），支持交互操作",
     "【前端】 1. Mind Elixir 5 渲染 ✅\n"
     "【前端】 2. 树视图\u2194脑图一键切换 ✅\n"
     "【前端】 3. 暗色主题适配（分层色/分支轮询/\u201c默认\u201d虚线标识/自适应缩放）✅\n"
     "【前端】 4. 脑图内拖拽调整层级 ✅\n"
     "【前端】 5. 脑图节点双击重命名 ✅\n"
     "【前端】 6. 连线投递到脑图\u201c默认\u201d区 ✅\n"
     "【后端】 1. 脑图数据查询 API（与指标树同源）\U0001F534\n"
     "【后端】 2. 脑图变更同步 API \U0001F534",
     "MindMapWrapper.tsx\nmindMapAdapter.ts\nIndicatorTreePanel.tsx"],
    ["指标管理", "标签树高级交互",
     "标签集树状层级展示与操作",
     "【前端】 1. 标签树层级展示 ✅\n"
     "【前端】 2. 标签选中/取消（Space 切换）✅\n"
     "【前端】 3. 标签节点增删 ✅\n"
     "【后端】 1. 标签集层级查询 API \U0001F534\n"
     "【后端】 2. 标签节点 CRUD API \U0001F534",
     "TagSetPanel.tsx\nTagPill.tsx\nTagCloud.tsx"],
    ["指标管理", "规则树高级交互",
     "规则配置树状层级展示与参数配置",
     "【前端】 1. 规则树层级展示 ✅\n"
     "【前端】 2. 规则选中与关联 ✅\n"
     "【前端】 3. 规则参数配置（阈值等）✅\n"
     "【后端】 1. 规则树层级查询 API \U0001F534\n"
     "【后端】 2. 规则参数配置存储 API \U0001F534",
     "RulePanel.tsx\nParameterDrawer.tsx\nParameterFields.tsx"],
    ["指标管理", "指标配置",
     "选中指标后在侧边栏/命令面板配置关联标签与规则",
     "【前端】 1. 已关联标签/规则侧边栏展示 ✅\n"
     "【前端】 2. 标签关联添加/移除 ✅\n"
     "【前端】 3. 规则关联添加/移除 ✅\n"
     "【前端】 4. Ctrl+K 命令面板快速关联 ✅\n"
     "【后端】 1. 指标关联关系查询 API \U0001F534\n"
     "【后端】 2. 指标关联关系更新 API \U0001F534",
     "IndicatorAttachmentPage.tsx\nAttachmentCommandPalette.tsx\nattachmentStore.ts"],

    # ── 血缘画布 ──
    ["血缘画布", "血缘关系可视化+实例管理",
     "指标间血缘关系可视化画布展示与关系实例增删改查",
     "【前端】 1. 节点+连线画布可视化 ✅\n"
     "【前端】 2. 节点聚焦高亮与路径追踪 ✅\n"
     "【前端】 3. 血缘关系实例增删改查 ✅\n"
     "【前端】 4. 实例筛选（AI推荐/人类/全部）✅\n"
     "【后端】 1. 血缘关系实例 CRUD API \U0001F534\n"
     "【后端】 2. 血缘关系类型定义查询 API \U0001F534",
     "LineageCanvasPage.tsx\ngetNodeStyle.ts\nAiRecommendationList.tsx"],

    # ── 报告管理 ──
    ["报告管理", "报告计划管理",
     "报告计划创建、编辑、删除与自动排程控制",
     "【前端】 1. 报告计划 CRUD 列表 ✅\n"
     "【前端】 2. 自动排程开关控制 ✅\n"
     "【后端】 1. 报告计划 CRUD API \U0001F534\n"
     "【后端】 2. 排程调度（定时触发）\U0001F534",
     "ReportManagementPage.tsx\nReportPlanDialog.tsx\nreportModel.ts"],
    ["报告管理", "报告生成与历史",
     "按计划手动生成报告，历史报告查看/筛选/版本回溯",
     "【前端】 1. 生成报告（按计划手动 + 按历史重新执行）✅\n"
     "【前端】 2. 报告历史列表与筛选 ✅\n"
     "【前端】 3. 报告查看 ✅\n"
     "【前端】 4. 版本管理与回溯 ✅\n"
     "【后端】 1. 报告生成引擎\u2b50（模板渲染+数据拉取+脚本管理）\U0001F534\n"
     "【后端】 2. 报告历史与版本存储 API \U0001F534",
     "ReportManagementPage.tsx\nReportHistoryPage.tsx\nReportDetailPage.tsx\nmockReportData.ts"],
    ["报告管理", "报告模板管理",
     "模板创建、编辑、章节编排及 AI 生成章节草案",
     "【前端】 1. 模板创建与维护 ✅\n"
     "【前端】 2. 模板章节编排排序 ✅\n"
     "【前端】 3. AI 生成模板章节草案 ✅\n"
     "【后端】 1. 报告模板 CRUD API \U0001F534\n"
     "【后端】 2. AI 生成模板章节 API \U0001F534",
     "ReportTemplatesPage.tsx\nreportTemplateModel.ts\nreportTemplateStorage.ts"],

    # ── 规则管理 ──
    ["规则管理", "规则树高级交互",
     "规则分类树层级展示与拖拽排序管理",
     "【前端】 1. 分类树展开/折叠 ✅\n"
     "【前端】 2. 分类节点新增/删除 ✅\n"
     "【前端】 3. 分类节点拖拽排序 ✅\n"
     "【后端】 1. 规则分类树 CRUD API \U0001F534",
     "NocRulePage.tsx"],
    ["规则管理", "规则实例增删改查",
     "规则实例列表展示、CRUD、JSON 参数配置与父规则继承",
     "【前端】 1. 规则列表与搜索筛选 ✅\n"
     "【前端】 2. 规则创建与编辑 ✅\n"
     "【前端】 3. JSON 参数配置编辑器 ✅\n"
     "【前端】 4. 父规则继承关联 ✅\n"
     "【后端】 1. 规则实例 CRUD API \U0001F534\n"
     "【后端】 2. 规则参数存储 API \U0001F534",
     "NocRulePage.tsx\nRuleSummaryBadge.tsx"],

    # ── 关联关系管理 ──
    ["关联关系管理", "关联关系实例管理",
     "血缘关系实例 ID 列表管理、CRUD、引用详情与变更日志",
     "【前端】 1. 列表筛选（方向/来源类型/部门/时间）✅\n"
     "【前端】 2. 关系实例增删改查 ✅\n"
     "【前端】 3. 展开详情查看引用关系 ✅\n"
     "【前端】 4. 变更时间线 ✅\n"
     "【后端】 1. 关系实例 CRUD API \U0001F534\n"
     "【后端】 2. 变更日志查询 API \U0001F534",
     "LinkRelationManagePage.tsx\nLinkRelationFormDialog.tsx\nChangeTimeline.tsx"],
    ["关联关系管理", "AI 推荐关系管理",
     "AI 推荐候选血缘关系列表，置信度筛选与一键应用",
     "【前端】 1. AI 推荐列表与置信度筛选 ✅\n"
     "【前端】 2. 勾选一键应用到血缘画布 ✅\n"
     "【后端】 1. AI 推荐关系查询 API \U0001F534",
     "AiRecommendationList.tsx\nlinkRelationModel.ts"],

    # ── 知识库管理 ──
    ["知识库管理", "知识库管理",
     "对接晓悟知识库 API，文档上传、查询与版本管理",
     "【后端】 1. 对接晓悟知识库 API（文档上传/查询/版本管理）\U0001F534",
     "KnowledgeUploadPage.tsx\nknowledgeBaseStorage.ts"],
]

# ========== Sheet 2: 工作量估算 ==========

estimate_data = [
    # 分类, 业务场景, 预计AI对话轮数, 复杂度, 风险说明

    ["基础设施", "Vue 项目脚手架+路由+菜单+布局", "2-3", "中", "Vite + Vue Router + Layout"],
    ["基础设施", "深色主题系统+Pinia 状态管理", "3-4", "中", "CSS变量体系迁移，Zustand\u2192Pinia"],
    ["基础设施", "UI组件库集成+通用组件移植", "2-3", "低", "Element Plus/Naive UI 替换 shadcn"],

    ["业务功能", "概览统计卡片", "1-2", "低", "纯展示组件"],
    ["业务功能", "快捷入口导航", "1", "低", "路由跳转卡片"],
    ["业务功能", "近期动态时间线", "1-2", "低", "列表+图标映射"],

    ["业务功能", "指标挂靠", "10-14", "高", "SVG实时连线跟随/hover检测/动画/持久连线/脑图投递区/多目标识别"],
    ["业务功能", "指标树高级交互", "6-8", "中", "树组件重构/inline编辑/拖拽排序/搜索定位"],
    ["业务功能", "脑图可视化+交互", "8-12", "高", "Mind Elixir 5 无Vue3封装，需手写wrapper"],
    ["业务功能", "标签树高级交互", "3-4", "低", "树+Space切换选中"],
    ["业务功能", "规则树高级交互", "3-4", "低", "树+参数侧边栏联动"],
    ["业务功能", "指标配置", "4-6", "中", "命令面板+侧边栏联动+批量关联"],

    ["业务功能", "血缘关系可视化+实例管理", "8-12", "高", "SVG画布自绘：节点布局/连线/拖拽/缩放/高亮"],
    ["业务功能", "报告计划管理", "3-4", "低", "CRUD表单+列表+开关"],
    ["业务功能", "报告生成与历史", "4-6", "中", "列表/筛选/分页/版本回溯"],
    ["业务功能", "报告模板管理", "4-6", "中", "CRUD+章节拖拽+AI生成交互"],
    ["业务功能", "规则树高级交互", "3-4", "低", "树展开折叠+拖拽排序"],
    ["业务功能", "规则实例增删改查", "4-6", "中", "列表+JSON编辑器+父规则继承"],
    ["业务功能", "关联关系实例管理", "4-6", "中", "表格CRUD+展开详情+变更时间线"],
    ["业务功能", "AI推荐关系管理", "3-5", "低", "列表筛选+勾选应用"],
    ["业务功能", "知识库管理", "1-2", "低", "纯后端API对接"],
]

# =================== 生成Excel ===================

wb = Workbook()

# ── 样式通用 ──
hfont = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
hfill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
halign = Alignment(horizontal="center", vertical="center", wrap_text=True)

bfont = Font(name="Microsoft YaHei", size=10)
balign = Alignment(vertical="top", wrap_text=True)
border = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

module_fills = {
    "首页/工作台": "EBF5FB", "指标管理": "FEF9E7", "血缘画布": "E8F8F5",
    "报告管理": "F4ECF7", "规则管理": "FDEDEC", "关联关系管理": "FDEBD0",
    "知识库管理": "EAECEE",
}
red_f = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
yellow_f = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
green_f = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
infra_f = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")


def fill_header(ws, headers):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hfont; c.fill = hfill; c.alignment = halign; c.border = border


# ── Sheet 1: 功能清单 ──
ws1 = wb.active
ws1.title = "功能清单"

s1_headers = ["模块", "业务场景", "描述", "子功能拆解（前端/后端+完成状态）", "当前 React 版核心 files"]
fill_header(ws1, s1_headers)

for ri, row in enumerate(data, 2):
    mod = row[0]
    mf = PatternFill(start_color=module_fills.get(mod, "FFFFFF"),
                     end_color=module_fills.get(mod, "FFFFFF"), fill_type="solid")
    for ci, v in enumerate(row, 1):
        c = ws1.cell(row=ri, column=ci, value=v)
        c.font = bfont; c.alignment = balign; c.border = border
        if ci == 1:
            c.fill = mf

ws1.column_dimensions["A"].width = 14
ws1.column_dimensions["B"].width = 22
ws1.column_dimensions["C"].width = 36
ws1.column_dimensions["D"].width = 66
ws1.column_dimensions["E"].width = 46
ws1.row_dimensions[1].height = 28
for r in range(2, len(data) + 2):
    ws1.row_dimensions[r].height = max(
        (ws1.cell(row=r, column=4).value or "").count("\n") * 16 + 20, 28)
ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:E{len(data)+1}"


# ── Sheet 2: 工作量估算 ──
ws2 = wb.create_sheet("AI重构工作量估算")

s2_headers = ["分类", "业务场景/基础设施", "预计 AI 对话轮数", "复杂度", "风险说明"]
fill_header(ws2, s2_headers)

for ri, row in enumerate(estimate_data, 2):
    for ci, v in enumerate(row, 1):
        c = ws2.cell(row=ri, column=ci, value=v)
        c.font = bfont; c.alignment = balign; c.border = border
        if ci == 1:
            c.fill = infra_f
        if ci == 4:
            if v == "高": c.fill = red_f
            elif v == "中": c.fill = yellow_f
            elif v == "低": c.fill = green_f

# 汇总行
tr = len(estimate_data) + 3

tc = ws2.cell(row=tr, column=1, value="合计")
tc.font = Font(name="Microsoft YaHei", size=11, bold=True, color="1F2937")
tc.alignment = Alignment(horizontal="right", vertical="center")

tc2 = ws2.cell(row=tr, column=2, value="基础设施 7-10 轮 + 业务场景 69-100 轮")
tc2.font = Font(name="Microsoft YaHei", size=11, bold=True, color="1F2937")
tc2.alignment = balign

tc3 = ws2.cell(row=tr, column=3, value="76-110 轮")
tc3.font = Font(name="Microsoft YaHei", size=11, bold=True, color="DC2626")

tc4 = ws2.cell(row=tr, column=4, value="约 4-6 周（1人全职）")
tc4.font = Font(name="Microsoft YaHei", size=11, bold=True, color="1F2937")
tc4.alignment = balign

ws2.column_dimensions["A"].width = 10
ws2.column_dimensions["B"].width = 26
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["D"].width = 10
ws2.column_dimensions["E"].width = 52
ws2.row_dimensions[1].height = 28
for r in range(2, tr + 1):
    ws2.row_dimensions[r].height = 22
ws2.row_dimensions[tr].height = 28
ws2.freeze_panes = "A2"

wb.save(out_path)
print(f"OK: {out_path}")
print(f"Sheet1: {len(data)} 行, Sheet2: {len(estimate_data)} 场景 + 汇总")
print(f"预估 AI 对话: 76-110 轮, 约 4-6 周（1人全职）")
